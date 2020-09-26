import json
import urllib.parse
import urllib.request
import datetime
import openpyxl
import pandas
import matplotlib.pyplot as plt


class Wallet:

    def __init__(self, wallet):
        self.wallet = wallet
        self.position = None
        for i in range(1, wallet.max_row+1):
            if wallet["B" + str(i)].value == 'latest':
                self.position = i
                break
        else:
            self.position = 3

        
        self.spent = dict()
        self.earned = dict()
        self._ids = 0
        self._ide = 0


    def add(self, earned:int, category:str):
        
        self.wallet["B" + str(self.position)] = ""
        self.wallet['D2'] = int(self.wallet['D2'].value) + earned
        self.wallet["D" + str(self.position)] = self.wallet['D2'].value
        self.wallet["C" +str(self.position)] = datetime.datetime.now().strftime("%Y/%m/%d, %H:%M:%S")
        self.wallet['E2'] = int(self.wallet['E2'].value) + earned
        self.wallet["E" + str(self.position)] = earned
        self.wallet["G" + str(self.position)] = category

        self.position += 1
        self.wallet["B" + str(self.position)] = "latest"


    def remove(self, spent:int, category:str):
        self.wallet["B" + str(self.position)] = ""
        self.wallet['D2'] = int(self.wallet['D2'].value) + spent
        self.wallet["D" + str(self.position)] = self.wallet['D2'].value
        self.wallet["C" +str(self.position)] = datetime.datetime.now().strftime("%Y/%m/%d, %H:%M:%S")
        self.wallet['F2'] = int(self.wallet['E2'].value) + spent
        self.wallet["F" + str(self.position)] = spent
        self.wallet["G" + str(self.position)] = category

        self.position += 1
        self.wallet["B" + str(self.position)] = "latest"

    def show_wallet(self):
        print("You have {} {} for total".format(self.wallet["D2"].value, self.wallet["A2"].value))

    def convert(self, amount:int, base_currency:str, foreign_currency:str):
        """Shows converted currency. Uses Euro Bank API"""
        currency_url = "https://api.exchangeratesapi.io/latest?base=" + base_currency
        result = None
        try:
            result = urllib.request.urlopen(currency_url)
            json_text = result.read().decode(encoding = "utf-8")

            converter = json.loads(json_text)
            return amount * converter["rates"][foreign_currency]
        except:
            pass

        finally:
            if result != None:
                result.close()
        


    def show_spent_earned(self):
        
        earned_data = ""
        spent_data = ""
        for e in range(3, self.position):
            temp_earned = self.wallet["E" + str(e)].value
            temp_spent = self.wallet["F" + str(e)].value
            if temp_earned != None:
                earned_data += "Earned {} {} in ".format(temp_earned, self.wallet["A2"].value) + str(self.wallet["C" + str(e)].value) + " " +\
                               str(self.wallet["G" + str(e)].value) + "\n"
            if temp_spent != None:
                spent_data += "Spent {} {} in ".format(temp_spent, self.wallet["A2"].value) + str(self.wallet["C" + str(e)].value) + " " + \
                              str(self.wallet["G" + str(e)].value) + "\n"

        print("-------Earned--------")
        print("You have earned {} {} for total".format(self.wallet["E2"].value, self.wallet["A2"].value))
        print(earned_data)
        print("-------Spent--------")
        print("You have spent {} {} for total".format(self.wallet["F2"].value, self.wallet["A2"].value))
        print(spent_data)
        



def load_wallet(wb, sheetname:str):
    """Opens and loads excel file sheet """
    sheet = wb[sheetname]
    return sheet
    


def create_wallets(total:int, currency: str):
    """Creates new workbook for wallet"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Wallet"
    sheet['A1'] = "Currency"
    sheet['A2'] = currency
    sheet['B2'] = "Total"
    sheet['C1'] = "Date"
    sheet['D1'] = "Wallet"
    sheet['E1'] = "Earned"
    sheet['F1'] = "Spent"
    sheet['G1'] = "Reason"
    sheet['D2'] = total
    sheet['E2'] = 0
    sheet['F2'] = 0
    return wb

def create_sheet(wb, sheetname:str, total:int, currency: str):
    """Creates new workbook for wallet"""
    wb.create_sheet(title=sheetname)
    sheet = wb[sheetname]
    sheet['A1'] = "Currency"
    sheet['A2'] = currency
    sheet['B2'] = "Total"
    sheet['C1'] = "Date"
    sheet['D1'] = "Wallet"
    sheet['E1'] = "Earned"
    sheet['F1'] = "Spent"
    sheet['G1'] = "Reason"
    sheet['D2'] = total
    sheet['E2'] = 0
    sheet['F2'] = 0
    return sheet

def perform_analysis_bar(sheet:str):
    """Reads excel file with pandas and extracts info to use in matplot.
    It shows the average spending per category"""
    df = pandas.read_excel("My_Wallets.xlsx", sheet_name=sheet, index_col = "Date", \
                           parse_dates=True, usecols=[2,4,5,6])
    avg_spending_cat = df.groupby("Reason").mean().sort_values(by="Spent")
    print(avg_spending_cat)
    fig, ax = plt.subplots()
    avg_spending_cat.plot(kind="barh", ax=ax, figsize=(17,6))
    ax.set_xlabel("Amount Spent", size=20)
    ax.set_ylabel("Category", size=20)
    plt.show()

def perform_analysis_line(sheet:str, date: str):
    df = pandas.read_excel("My_Wallets.xlsx", sheet_name=sheet, index_col = "Date", \
                           parse_dates=True, usecols=[2,4,5,6])
    avg_spending_cat = df.groupby("Reason").mean().sort_values(by="Spent")
    monthly_spending = df.resample(date).mean()
    # Plotting monthly_spending
    monthly_spending.plot(figsize=(17,6))
    plt.xlabel('Date', size=20)
    plt.ylabel('Spent', size=20)
    plt.title('Average Monthly Spending')
    plt.show()
    

def main_menu():
    running = True
    wallet = None
    wallets = None
    sheetname = None
    try:
        wallets = openpyxl.load_workbook("My_Wallets.xlsx")
    except:
        wallets = create_wallets(0, "USD")
    while running:
        print("Welcome to your Wallet")
        print("Press [C] to create wallet")
        print("Press [L] to load wallet")
        print("Press [A] to add earnings")
        print("Press [S] to add spendings")
        print("Press [Q] to show your wallet")
        print("Press [F] to convert your wallet to a foreign currency")
        print("Press [W] to show analysis")
        print("Press [E] to quit")
        print("**WALLET IS 0 USD AS DEFAULT**")
        print("Current selected wallet: " + str(sheetname))
        command = input("Command: ")
        print("--------------------------------------------------------")
        if command == "C":
            print("Creating Wallet...")
            name = input("Name of wallet: ")
            total = int(input("How much do you have? "))
            currency = input("What is the currency? ")
            sheetname = name
            try:
                w = create_sheet(wallets, name, total, currency)
                wallet = Wallet(w)
                print("Wallet has been created\n")
            except:
                print("Could not create wallet...")

        elif command == "L":
            print("Loading Wallet....")
            name = input("Name of wallet: ")
            sheetname = name
            try:
                w = load_wallet(wallets, name)
                wallet = Wallet(w)
            except:
                print("Not a valid name of wallet.")
                print("Write the specific name")
                
        elif command == "A":
            print("Adding...")
            amount_to_add = input("Amount: ")
            category = input("Category: ")
            try:
                wallet.add(int(amount_to_add), category)
                print("It has been added to your wallet\n")
            except:
                print("Could not add earning to wallet...")

            
        elif command == "S":
            print("Removing...")
            amount_to_remove = input("Amount: ")
            category = input("Category: ")
            try:
                wallet.remove(int(amount_to_remove), category)
                print("It has been deducted from your wallet\n")
            except:
                print("Could not add spending to wallet...")
            
        elif command == "Q":
            print("YOUR WALLET:")
            try:
                wallet.show_wallet()
                wallet.show_spent_earned()
            except:
                print("There are some error")
            
        elif command == "F":
            print("Converting currency")
            while True:
                amount = input("Do you want all your [wallet] or [specify] amount? ")
                base_currency = input("Base currency: ")
                foreign_currency = input("Foreign currency: ")
                if amount == "wallet":
                    converted = wallet.convert(wallet.wallet["D2"].value, base_currency, foreign_currency)
                    print("{} {} to {}".format(str(wallet.wallet["D2"].value), base_currency, foreign_currency))
                    print("It is {:.2f} {}\n".format(converted, foreign_currency))
                    break
                elif amount == "specify":
                    total = int(input("Amount to convert: "))
                    converted = wallet.convert(total, base_currency, foreign_currency)
                    print("{} {} to {}".format(str(total), base_currency, foreign_currency))
                    print("It is {:.2f} {}\n".format(converted, foreign_currency))
                    break
                else:
                    print("Invalid Command")
                    
        elif command == "E":
            running = False
            try:
                wallets.close()
                wallets.save("My_Wallets.xlsx")
            except:
                return

        elif command == "W":
            type_of_analysis = input("[Bar] or [Line]? ")
            if type_of_analysis == "Bar":
                try:
                    perform_analysis_bar(sheetname)
                except:
                    print("There was an error")
            elif type_of_analysis == "Line":
                time = input("[d]aily, [m]onthly or [y]early? ")
                time_range = None
                if time == "d":
                    time_range = "D"
                elif time == "m":
                    time_range = "M"
                elif time == "y":
                    time_range = "Y"
                else:
                    print("Please, select d, m, or y")
                    
                try:
                    perform_analysis_line(sheetname, time_range)
                except:
                    print("There was an error")
            else:
                print("Please write the correct method")
            
        else:
            print("Invalid Command\n")

        print("--------------------------------------------------------")









if __name__ == "__main__":
    main_menu()


    #https://towardsdatascience.com/analysing-your-finances-with-pandas-and-matplotlib-b587a8089bb2

    
